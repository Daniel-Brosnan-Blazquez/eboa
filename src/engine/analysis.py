"""
Data analysis using openpyxl for generating excel 2010 workbenches

Written by DEIMOS Space S.L. (dibb)

module gsdm
"""
# Import openpyxl functionalities
from openpyxl import Workbook
from openpyxl.styles import Font
from .engine import Engine

class Analysis():
    workbook = Workbook()
    engine = Engine()

    def generate_workbook_from_ddbb(self, filename):
        """
        """
        # Remove the sheet created by default
        self.workbook.remove(self.workbook.active)

        self._insert_dim_signatures()
        self._insert_sources()
        self._insert_sources_statuses()
        self._insert_events()
        self._insert_gauges()
        self._insert_events_keys()
        self._insert_events_links()
        self._insert_annotations()
        self._insert_annotations_cnfs()
        self._insert_explicit_references()
        self._insert_explicit_references_links()
        self._insert_explicit_references_groups()

        # Save the workbook into the file specified
        self.workbook.save(filename)

        return

    def _insert_dim_signatures(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("DIM Signatures")
        
        # Get data
        dim_signatures = self.engine.get_dim_signatures()
        data = [[i.dim_signature_id,i.dim_signature,i.dim_exec_name] for i in dim_signatures]
        
        # Insert headings into the worksheet
        ws.append(["Id", "Name", "Processor name"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return        

    def _insert_sources(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Sources")
        
        # Get data
        sources = self.engine.get_sources()
        data = [[str(i.processing_uuid),i.filename,i.validity_start,i.validity_stop,i.generation_time,i.ingestion_time,i.ingestion_duration,i.dim_exec_version,i.dim_signature_id] for i in sources]
        
        # Insert headings into the worksheet
        ws.append(["Id", "Name", "Validity start", "Validity stop", "Generation time", "Ingestion time", "Ingestion duration", "Version", "DIM signature id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return

    def _insert_sources_statuses(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Sources Statuses")
        
        # Get data
        sources_statuses = self.engine.get_sources_statuses()
        data = [[i.time_stamp,i.proc_status,str(i.processing_uuid)] for i in sources_statuses]
        
        # Insert headings into the worksheet
        ws.append(["Time stamp", "Status", "Source UUID"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return

    def _insert_events(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Events")
        
        # Get data
        events = self.engine.get_events()
        data = [[str(i.event_uuid),i.start,i.stop,i.generation_time,i.ingestion_time,i.visible, i.gauge_id,i.explicit_ref_id,str(i.processing_uuid)] for i in events]
        
        # Insert headings into the worksheet
        ws.append(["Event UUID", "start", "stop", "Generation Time", "Ingestion time", "Visible", "Gauge id", "Explicit reference id", "Processing id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return

    def _insert_gauges(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Gauges")
        
        # Get data
        gauges = self.engine.get_gauges()
        data = [[i.gauge_id,i.system,i.name,i.dim_signature_id] for i in gauges]
        
        # Insert headings into the worksheet
        ws.append(["Gauge id", "System", "Name", "DIM signature id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return

    def _insert_events_keys(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Event Keys")
        
        # Get data
        event_keys = self.engine.get_event_keys()
        data = [[i.event_key,i.generation_time,i.visible,str(i.event_uuid),i.dim_signature_id] for i in event_keys]
        
        # Insert headings into the worksheet
        ws.append(["Event key", "Generation time", "Visible", "Event UUID", "DIM signature id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return

    def _insert_events_links(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Event Links")
        
        # Get data
        event_links = self.engine.get_event_links()
        data = [[str(i.event_uuid_link),i.name,str(i.event_uuid)] for i in event_links]
        
        # Insert headings into the worksheet
        ws.append(["Event UUID", "Link name", "Event uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return

    def _insert_annotations(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Annotations")
        
        # Get data
        annotations = self.engine.get_annotations()
        data = [[str(i.annotation_uuid),i.generation_time,i.ingestion_time,i.visible,i.explicit_ref_id,str(i.processing_uuid),i.annotation_cnf_id] for i in annotations]
        
        # Insert headings into the worksheet
        ws.append(["Annotation UUID", "Generation time", "Ingestion time", "Visible", "Explicit reference id", "Source UUID", "Annotation configuration id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return

    def _insert_annotations_cnfs(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Annotations configurations")
        
        # Get data
        annotations = self.engine.get_annotations_configurations()
        data = [[i.annotation_cnf_id,i.system,i.name,i.dim_signature_id] for i in annotations]
        
        # Insert headings into the worksheet
        ws.append(["Annotation configuration id", "System", "Name", "DIM signature id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return


    def _insert_explicit_references(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Explicit references")
        
        # Get data
        explicit_refs = self.engine.get_explicit_references()
        data = [[i.explicit_ref_id,i.ingestion_time,i.explicit_ref,i.expl_ref_cnf_id] for i in explicit_refs]
        
        # Insert headings into the worksheet
        ws.append(["Explicit reference id", "Ingestion time", "Explicit reference", "Group"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return

    def _insert_explicit_references_links(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Explicit references links")
        
        # Get data
        explicit_refs_links = self.engine.get_explicit_references_links()
        data = [[i.explicit_ref_id_link,i.name,i.explicit_ref_id] for i in explicit_refs_links]
        
        # Insert headings into the worksheet
        ws.append(["Explicit reference id", "Link name", "Explicit reference linked"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return

    def _insert_explicit_references_groups(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("Explicit references groups")
        
        # Get data
        explicit_refs_groups = self.engine.get_explicit_references_groups()
        data = [[i.expl_ref_cnf_id,i.name] for i in explicit_refs_groups]
        
        # Insert headings into the worksheet
        ws.append(["Explicit reference group id", "Group name"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(bold="True")

        return


