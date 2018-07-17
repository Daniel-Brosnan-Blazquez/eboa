"""
Data analysis using openpyxl for generating excel 2010 workbenches

Written by DEIMOS Space S.L. (dibb)

module gsdm
"""
# Import openpyxl functionalities
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from .query import Query
from .plotting import generate_gantt

# Import GEOalchemy entities
from geoalchemy2.shape import to_shape

# Import matplotlib
from openpyxl.drawing.image import Image

# Import python utilities
import numpy
import os

class Analysis():
    def __init__(self):
        self.workbook = Workbook()
        self.query_gsdm = Query()

    def generate_workbook_from_ddbb(self, name):
        """
        """
        self.name = name
        self.files_to_remove = []
        # Remove the sheet created by default
        self.workbook.remove(self.workbook.active)

        self._insert_dim_signatures()
        self._insert_sources()
        self._insert_sources_statuses()
        self._insert_events()
        self._insert_gauges()
        self._insert_events_keys()
        self._insert_events_links()
        self._insert_annotations()
        self._insert_annotations_cnfs()
        self._insert_explicit_references()
        self._insert_explicit_references_links()
        self._insert_explicit_references_groups()
        self._insert_event_booleans()
        self._insert_event_texts()
        self._insert_event_doubles()
        self._insert_event_timestamps()
        self._insert_event_objects()
        self._insert_event_geometries()
        self._insert_annotation_booleans()
        self._insert_annotation_texts()
        self._insert_annotation_doubles()
        self._insert_annotation_timestamps()
        self._insert_annotation_objects()
        self._insert_annotation_geometries()

        # Save the workbook into the file specified
        self.workbook.save(name)

        for file in self.files_to_remove:
            os.remove(file)
        # end for

        return

    def _adjust_column_width(self, ws):
        """
        """
        for column in ws.columns:
            length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column].width = length + 2
        # end for
        return

    def _insert_dim_signatures(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("dim_signature_tb")
        
        # Get data
        dim_signatures = self.query_gsdm.get_dim_signatures()
        data = [[i.dim_signature_id,i.dim_signature,i.dim_exec_name] for i in dim_signatures]
        
        # Insert headings into the worksheet
        ws.append(["dim_signature_id", "dim_signature", "dim_exec_name"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")
        
        self._adjust_column_width(ws)

        return        

    def _insert_sources(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("dim_processing_tb")
        
        # Get data
        sources = self.query_gsdm.get_sources()
        data = [[str(i.processing_uuid),i.name,i.validity_start,i.validity_stop,i.generation_time,i.ingestion_time,i.ingestion_duration,i.dim_exec_version,i.dim_signature_id] for i in sources]
        
        # Insert headings into the worksheet
        ws.append(["processing_uuid", "name", "validity_start", "validity_stop", "generation_time", "ingestion_time", "ingestion_duration", "dim_exec_version", "dim_signature_id"])

        # Insert data into the worksheet and prepare it for creating the gantt
        y_labels_gantt = []
        data_gantt = []
        for row in data:
            ws.append(row)
            validity_start = row[2]
            validity_stop = row[3]
            if validity_start != None and validity_stop != None:
                y_labels_gantt.append(row[1])
                data_gantt.append({"start": validity_start, "stop": validity_stop, "color": "blue", "text": row[4]})
            # end if
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        filename = generate_gantt(y_labels_gantt, data_gantt)
        self.files_to_remove.append(filename)
        img = Image(filename)
        ws.add_image(img, "A" + str(len(data)+4))

        return

    def _insert_sources_statuses(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("dim_processing_status_tb")
        
        # Get data
        sources_statuses = self.query_gsdm.get_sources_statuses()
        data = [[i.time_stamp,i.proc_status,str(i.processing_uuid)] for i in sources_statuses]
        
        # Insert headings into the worksheet
        ws.append(["time_stamp", "proc_status", "processing_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_events(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("event_tb")
        
        # Get data
        events = self.query_gsdm.get_events()
        data = [[str(i.event_uuid),i.start,i.stop,i.ingestion_time,i.visible, i.gauge_id,i.explicit_ref_id,str(i.processing_uuid)] for i in events]
        
        # Insert headings into the worksheet
        ws.append(["event_uuid", "start", "stop", "ingestion_time", "visible", "gauge_id", "explicit_ref_id", "processing_uuid"])

        # Insert data into the worksheet and prepare it for creating the gantt
        y_labels_gantt = []
        data_gantt = []
        for row in data:
            ws.append(row)
            start = row[1]
            stop = row[2]
            if start != None and stop != None:
                y_labels_gantt.append(row[0])
                data_gantt.append({"start": start, "stop": stop, "color": "blue", "text": row[3]})
            # end if
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        filename = generate_gantt(y_labels_gantt, data_gantt)
        self.files_to_remove.append(filename)
        img = Image(filename)
        ws.add_image(img, "A" + str(len(data)+4))

        return

    def _insert_gauges(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("gauge_cnf_tb")
        
        # Get data
        gauges = self.query_gsdm.get_gauges()
        data = [[i.gauge_id,i.system,i.name,i.dim_signature_id] for i in gauges]
        
        # Insert headings into the worksheet
        ws.append(["gauge_id", "system", "name", "dim_signature_id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_events_keys(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("event_keys_tb")
        
        # Get data
        event_keys = self.query_gsdm.get_event_keys()
        data = [[i.event_key,i.visible,str(i.event_uuid),i.dim_signature_id] for i in event_keys]
        
        # Insert headings into the worksheet
        ws.append(["event_key", "visible", "event_uuid", "dim_signature_id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_events_links(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("event_links_tb")
        
        # Get data
        event_links = self.query_gsdm.get_event_links()
        data = [[str(i.event_uuid_link),i.name,str(i.event_uuid)] for i in event_links]
        
        # Insert headings into the worksheet
        ws.append(["event_uuid", "name", "event_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_annotations(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("annot_tb")
        
        # Get data
        annotations = self.query_gsdm.get_annotations()
        data = [[str(i.annotation_uuid),i.ingestion_time,i.visible,i.explicit_ref_id,str(i.processing_uuid),i.annotation_cnf_id] for i in annotations]
        
        # Insert headings into the worksheet
        ws.append(["annotation_uuid", "ingestion_time", "visible", "explicit_ref_id", "processing_uuid", "annotation_cnf_id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_annotations_cnfs(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("annot_cnf_tb")
        
        # Get data
        annotations = self.query_gsdm.get_annotations_configurations()
        data = [[i.annotation_cnf_id,i.system,i.name,i.dim_signature_id] for i in annotations]
        
        # Insert headings into the worksheet
        ws.append(["annotation_cnf_id", "system", "name", "dim_signature_id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return


    def _insert_explicit_references(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("explicit_ref_tb")
        
        # Get data
        explicit_refs = self.query_gsdm.get_explicit_references()
        data = [[i.explicit_ref_id,i.ingestion_time,i.explicit_ref,i.expl_ref_cnf_id] for i in explicit_refs]
        
        # Insert headings into the worksheet
        ws.append(["explicit_ref_id", "ingestion_time", "explicit_ref", "expl_ref_cnf_id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_explicit_references_links(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("explicit_ref_links_tb")
        
        # Get data
        explicit_refs_links = self.query_gsdm.get_explicit_references_links()
        data = [[i.explicit_ref_id_link,i.name,i.explicit_ref_id] for i in explicit_refs_links]
        
        # Insert headings into the worksheet
        ws.append(["explicit_ref_id_link", "name", "explicit_ref_id"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_explicit_references_groups(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("explicit_ref_cnf_tb")
        
        # Get data
        explicit_refs_groups = self.query_gsdm.get_explicit_references_groups()
        data = [[i.expl_ref_cnf_id,i.name] for i in explicit_refs_groups]
        
        # Insert headings into the worksheet
        ws.append(["expl_ref_cnf_id", "name"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_event_booleans(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("event_boolean_tb")
        
        # Get data
        values = self.query_gsdm.get_event_booleans()
        data = [[i.name,i.value,i.level_position,i.parent_level,i.parent_position,str(i.event_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "value", "level_position", "parent_level", "parent_position", "event_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_event_texts(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("event_text_tb")
        
        # Get data
        values = self.query_gsdm.get_event_texts()
        data = [[i.name,i.value,i.level_position,i.parent_level,i.parent_position,str(i.event_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "value", "level_position", "parent_level", "parent_position", "event_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_event_doubles(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("event_double_tb")
        
        # Get data
        values = self.query_gsdm.get_event_doubles()
        data = [[i.name,i.value,i.level_position,i.parent_level,i.parent_position,str(i.event_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "value", "level_position", "parent_level", "parent_position", "event_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_event_timestamps(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("event_timestamp_tb")
        
        # Get data
        values = self.query_gsdm.get_event_timestamps()
        data = [[i.name,i.value,i.level_position,i.parent_level,i.parent_position,str(i.event_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "value", "level_position", "parent_level", "parent_position", "event_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_event_objects(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("event_object_tb")
        
        # Get data
        values = self.query_gsdm.get_event_objects()
        data = [[i.name,i.level_position,i.parent_level,i.parent_position,str(i.event_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "level_position", "parent_level", "parent_position", "event_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_event_geometries(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("event_geometry_tb")
        
        # Get data
        values = self.query_gsdm.get_event_geometries()
        data = [[i.name,to_shape(i.value).to_wkt(),i.level_position,i.parent_level,i.parent_position,str(i.event_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "value", "level_position", "parent_level", "parent_position", "event_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_annotation_booleans(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("annotation_boolean_tb")
        
        # Get data
        values = self.query_gsdm.get_annotation_booleans()
        data = [[i.name,i.value,i.level_position,i.parent_level,i.parent_position,str(i.annotation_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "value", "level_position", "parent_level", "parent_position", "annotation_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_annotation_texts(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("annotation_text_tb")
        
        # Get data
        values = self.query_gsdm.get_annotation_texts()
        data = [[i.name,i.value,i.level_position,i.parent_level,i.parent_position,str(i.annotation_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "value", "level_position", "parent_level", "parent_position", "annotation_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_annotation_doubles(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("annotation_double_tb")
        
        # Get data
        values = self.query_gsdm.get_annotation_doubles()
        data = [[i.name,i.value,i.level_position,i.parent_level,i.parent_position,str(i.annotation_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "value", "level_position", "parent_level", "parent_position", "annotation_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_annotation_timestamps(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("annotation_timestamp_tb")
        
        # Get data
        values = self.query_gsdm.get_annotation_timestamps()
        data = [[i.name,i.value,i.level_position,i.parent_level,i.parent_position,str(i.annotation_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "value", "level_position", "parent_level", "parent_position", "annotation_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_annotation_objects(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("annotation_object_tb")
        
        # Get data
        values = self.query_gsdm.get_annotation_objects()
        data = [[i.name,i.level_position,i.parent_level,i.parent_position,str(i.annotation_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "level_position", "parent_level", "parent_position", "annotation_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return

    def _insert_annotation_geometries(self):
        """
        """
        # Create worksheet
        ws = self.workbook.create_sheet("annotation_geometry_tb")
        
        # Get data
        values = self.query_gsdm.get_annotation_geometries()
        data = [[i.name,to_shape(i.value).to_wkt(),i.level_position,i.parent_level,i.parent_position,str(i.annotation_uuid)] for i in values]
        
        # Insert headings into the worksheet
        ws.append(["name", "value", "level_position", "parent_level", "parent_position", "annotation_uuid"])

        # Insert data into the worksheet
        for row in data:
            ws.append(row)
        # end for

        # Applying styles
        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        self._adjust_column_width(ws)

        return




