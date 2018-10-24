"""
Test: ingest xml test

Written by DEIMOS Space S.L. (dibb)

module eboa
"""
# Import python utilities
import os
import sys
import argparse
import re

# Import eboa components
import eboa.engine.engine as engine
from eboa.engine.query import Query

# Import openpyxl functionalities
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# Import matplotlib
from openpyxl.drawing.image import Image

# Import auxiliary functions
from eboa.engine.functions import is_datetime
from eboa.analysis.functions import adjust_column_width

# Import logging
from eboa.logging import Log

logging = Log()
logger = logging.logger
missions = ["S2A", "S2B"]

def generate_imaging_analysis(workbook, query, begin, end):

    events = query.get_linked_events_join(gauge_name_like = {"str": "CUT_IMAGING%", "op": "like"}, start_filters = [{"date": end, "op": "<"}], stop_filters = [{"date": begin, "op": ">"}], link_names = {"list": ["RECORD_OPERATION", "COMPLETE_IMAGING_OPERATION"], "op": "in"})

    imaging_events_and_linked = events["linked_events"] + events["prime_events"]

    events = query.get_linked_events(event_uuids = [event.event_uuid for event in imaging_events_and_linked], start_filters = [{"date": end, "op": "<"}], stop_filters = [{"date": begin, "op": ">"}], link_names = {"list": ["TIME_CORRECTION"], "op": "in"}, return_prime_events = False)

    corrected_imaging_events = events["linked_events"]

    imaging_events_and_linked.sort(key=lambda k: k.__dict__["start"])

    # Imaging plan
    ws = workbook.create_sheet("Imaging plan")

    # Insert headings into the worksheet
    ws.append(["Satellite", "Orbit", "Gauge", "Start", "Stop", "Duration (m)", "Parameters", "Status", "Event uuid", "Ingestion time", "NPPF file", "DIM version"])

    # String patterns
    record_parameters_pattern = re.compile(".*scn_dup$")
    record_pattern = re.compile("^RECORD.*")
    cut_imaging_pattern = re.compile("^CUT_IMAGING.*")
    imaging_pattern = re.compile("^IMAGING.*")

    # Insert data into the worksheet
    for event in imaging_events_and_linked:
        parameters_text = ""
        if record_pattern.match(event.gauge.name):
            parameters = [parameter for parameter in event.eventDoubles if record_parameters_pattern.match(parameter.name)]
            i = 1
            for parameter in parameters:
                parameters_text += parameter.name + ": " + str(int(parameter.value))
                if i < len(parameters):
                    parameters_text += ", "
                # end if
                i += 1
            # end for
        # end if
        corrected_event_uuid = [link.event_uuid_link for link in event.eventLinks if link.name == "TIME_CORRECTION"]
        status = "TIME_CORRECTED"
        if len(corrected_event_uuid) > 0:
            corrected_event = [event for event in corrected_imaging_events if event.event_uuid == corrected_event_uuid[0]]
            start = corrected_event[0].start
            stop = corrected_event[0].stop
            status = [obj.value for obj in corrected_event[0].eventTexts if obj.name == "status_correction"][0]
        else:
            status = "TIME_NOT_CORRECTED"
            start = event.start
            stop = event.stop
        # end if
        orbit = [obj.value for obj in event.eventDoubles if obj.name == "start_orbit"][0]
        ws.append([event.gauge.system, str(int(orbit)), event.gauge.name, start, stop,  format((stop - start).total_seconds() / 60, ".3f"), parameters_text, status, str(event.event_uuid), event.ingestion_time, event.source.name, event.source.dim_exec_version])
    # end for

    # Applying styles
    blue_fill = PatternFill(start_color="00EAFF", end_color="00EAFF", fill_type="solid")
    green_fill = PatternFill(start_color="96EA82", end_color="96EA82", fill_type="solid")
    orange_fill = PatternFill(start_color="D9CAAC", end_color="D9CAAC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FDFF7B", end_color="FDFF7B", fill_type="solid")
    purple_fill = PatternFill(start_color="D57BFF", end_color="D57BFF", fill_type="solid")

    mission_colors = [yellow_fill, purple_fill]
    assigned_colors_mission = {}
    i = 0
    for mission in missions:
        assigned_colors_mission[mission] = mission_colors[i]
        i += 1
    # end for

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    row = ws.row_dimensions[1]
    row.font = Font(name="mono",bold="True")

    # Color record rows
    record_rows = [row for row in ws.rows if record_pattern.match(row[2].value)]
    for row in record_rows:
        for cell in row:
            cell.fill = blue_fill
            cell.border = thin_border
        # end for
        row[0].fill = assigned_colors_mission[row[0].value]
    # end for

    # Color cut imaging rows
    cut_imaging_rows = [row for row in ws.rows if cut_imaging_pattern.match(row[2].value)]
    for row in cut_imaging_rows:
        for cell in row:
            cell.fill = green_fill
            cell.border = thin_border
        # end for
        row[0].fill = assigned_colors_mission[row[0].value]
    # end for

    # Color cut imaging rows
    imaging_rows = [row for row in ws.rows if imaging_pattern.match(row[2].value)]
    for row in imaging_rows:
        for cell in row:
            cell.fill = orange_fill
            cell.border = thin_border
        # end for
        row[0].fill = assigned_colors_mission[row[0].value]
    # end for

    # Freeze first row
    ws.freeze_panes = "A2"

    # Adjust column widths
    adjust_column_width(ws)

    # Statistics
    imaging_durations = {}
    total_imaging_durations = 0
    orbits = []
    for mission in missions:
        ws = workbook.create_sheet("{} imaging statistics".format(mission))

        row = ws.row_dimensions[1]
        row.font = Font(name="mono",bold="True")

        # Insert headings into the worksheet
        ws.append(["Imaging mode", "Duration (m)", "", "Total (m)", "Net average load (m)"])

        # Extract statistics
        mission_imaging_durations = {}
        mission_total_imaging_durations = 0
        mission_orbits = {}
        for event in imaging_events_and_linked:
            if imaging_pattern.match(event.gauge.name) and event.gauge.system == mission:
                if not event.gauge.name in mission_imaging_durations:
                    mission_imaging_durations[event.gauge.name] = 0
                # end if
                if not event.gauge.name in imaging_durations:
                    imaging_durations[event.gauge.name] = 0
                # end if
                event_duration = float(format((event.stop - event.start).total_seconds() / 60, ".3f"))
                mission_imaging_durations[event.gauge.name] += event_duration
                imaging_durations[event.gauge.name] += event_duration
                mission_total_imaging_durations += event_duration
                total_imaging_durations += event_duration
                start_orbit = [record.value for record in event.eventDoubles if record.name == "start_orbit"][0]
                if not start_orbit in mission_orbits:
                    mission_orbits[start_orbit] = start_orbit
                    orbits.append(start_orbit)
                # end if
            # end if
        # end for

        for imaging in mission_imaging_durations:
            ws.append([imaging, mission_imaging_durations[imaging]])
        # end for

        ws["D2"] = mission_total_imaging_durations
        if len(mission_orbits.keys()) == 0:
            ws["E2"] = 0
        else:
            ws["E2"] = float(format(mission_total_imaging_durations / len(mission_orbits.keys()), ".3f"))

        # Adjust column widths
        adjust_column_width(ws)

    # end for

    ws = workbook.create_sheet("All mission imaging statistics")

    row = ws.row_dimensions[1]
    row.font = Font(name="mono",bold="True")
    
    # Insert headings into the worksheet
    ws.append(["Imaging mode", "Duration (m)", "", "Total (m)", "Net average load (m)"])

    for imaging in imaging_durations:
        ws.append([imaging, imaging_durations[imaging]])
    # end for

    ws["D2"] = total_imaging_durations
    if len(orbits) == 0:
        ws["E2"] = 0
    else:
        ws["E2"] = float(format(total_imaging_durations / len(orbits), ".3f"))
    # end if
    
    # Playback information
    ws = workbook.create_sheet("Downlink plan")
    playback_events = query.get_events_join(gauge_name_like = {"str": "PLAYBACK_TYPE%", "op": "like"}, start_filters = [{"date": end, "op": "<"}], stop_filters = [{"date": begin, "op": ">"}])

    playback_event_uuids = [event.event_uuid for event in playback_events]

    events = query.get_linked_events(event_uuids = playback_event_uuids, start_filters = [{"date": end, "op": "<"}], stop_filters = [{"date": begin, "op": ">"}], link_names = {"list": ["TIME_CORRECTION"], "op": "in"}, return_prime_events = False)

    corrected_playback_events = events["linked_events"]

    events = query.get_linked_events(event_uuids = playback_event_uuids, start_filters = [{"date": end, "op": "<"}], stop_filters = [{"date": begin, "op": ">"}], link_names = {"list": ["PLAYBACK_MEAN"], "op": "in"}, return_prime_events = False)

    playback_mean_events = events["linked_events"]

    events = query.get_linked_events(event_uuids = playback_event_uuids, start_filters = [{"date": end, "op": "<"}], stop_filters = [{"date": begin, "op": ">"}], link_names = {"list": ["DFEP_SCHEDULE"], "op": "in"}, return_prime_events = False)

    dfep_schedule_events = events["linked_events"]

    events = query.get_linked_events(event_uuids = playback_event_uuids, start_filters = [{"date": end, "op": "<"}], stop_filters = [{"date": begin, "op": ">"}], link_names = {"list": ["STATION_SCHEDULE"], "op": "in"}, return_prime_events = False)

    station_schedule_events = events["linked_events"]

    playback_events.sort(key=lambda k: k.__dict__["start"])

    # Insert headings into the worksheet
    ws.append(["Satellite", "Orbit", "Station", "Downlink Type", "Start", "Stop", "Duration (m)", "Parameters", "Status", "Event uuid", "Ingestion time", "NPPF file", "DIM version"])

    # Insert data into the worksheet
    for event in playback_events:
        parameters_text = ""
        parameters_object = [obj for obj in event.eventObjects if obj.name == "parameters"][0]
        parameters = [obj for obj in event.get_values() if obj.parent_level == parameters_object.parent_level + 1 and obj.parent_position == parameters_object.level_position]
        i = 1
        for parameter in parameters:
            parameters_text += parameter.name + ": " + str(int(parameter.value))
            if i < len(parameters):
                parameters_text += ", "
            # end if
            i += 1
        # end for

        corrected_event_uuid = [link.event_uuid_link for link in event.eventLinks if link.name == "TIME_CORRECTION"]
        status = "TIME_CORRECTED"
        if len(corrected_event_uuid) > 0:
            corrected_event = [event for event in corrected_playback_events if event.event_uuid == corrected_event_uuid[0]]
            start = corrected_event[0].start
            stop = corrected_event[0].stop
            status = [obj.value for obj in corrected_event[0].eventTexts if obj.name == "status_correction"][0]
        else:
            status = "TIME_NOT_CORRECTED"
            start = event.start
            stop = event.stop
        # end if
        orbit = [obj.value for obj in event.eventDoubles if obj.name == "start_orbit"][0]
        playback_mean_uuid = [link.event_uuid_link for link in event.eventLinks if link.name == "PLAYBACK_MEAN"]
        playback_mean_event = [event for event in playback_mean_events if event.event_uuid == playback_mean_uuid[0]][0]

        station = "CGSX"
        if playback_mean_event.gauge.name == "PLAYBACK_MEAN_XBAND":
            dfep_schedule_uuid = [link.event_uuid_link for link in event.eventLinks if link.name == "DFEP_SCHEDULE"]
            station_schedule_uuid = [link.event_uuid_link for link in event.eventLinks if link.name == "STATION_SCHEDULE"]
            if len(dfep_schedule_uuid) > 0:
                dfep_schedule_event = [event for event in dfep_schedule_events if event.event_uuid == dfep_schedule_uuid[0]]
                station = [obj.value for obj in dfep_schedule_event[0].eventTexts if obj.name == "station"][0]
            elif len(station_schedule_uuid) > 0:
                station_schedule_event = [event for event in station_schedule_events if event.event_uuid == station_schedule_uuid[0]]
                station = [obj.value for obj in station_schedule_event[0].eventTexts if obj.name == "station"][0]
            # end if
        else:
            station = "EDRS"
        # end if
        ws.append([event.gauge.system, str(int(orbit)), station, event.gauge.name.replace("PLAYBACK_TYPE_", ""), start, stop,  format((stop - start).total_seconds() / 60, ".3f"), parameters_text, status, str(event.event_uuid), event.ingestion_time, event.source.name, event.source.dim_exec_version])
    # end for


    # Adjust column widths
    adjust_column_width(ws)

    return

def generate_analysis(file_path, begin, end):
    """
    Method to generate the specific analysis into the file received as
    argument and following the period between the begin and end dates
    
    :param file_path: path to the file where the analysis is going to be stored
    :type file_path: str
    :param begin: start date of the period
    :type being: date
    :param end: stop date of the period
    :type end: date
    """
    
    query = Query()
    workbook = Workbook()
    workbook.remove(workbook.active)

    # Generate imaging analysis
    generate_imaging_analysis(workbook, query, begin, end)

    # Save spreadsheet
    workbook.save(file_path)

    return

if __name__ == "__main__":
    args_parser = argparse.ArgumentParser(description="Generate mission planning analysis.")
    args_parser.add_argument("-f", dest="file_path", type=str, nargs=1,
                             help="path to the file where to save the generated spreadsheet", required=True)
    args_parser.add_argument("-b", dest="begin", type=str, nargs=1,
                             help="begin date for the query period", required=True)
    args_parser.add_argument("-e", dest="end", type=str, nargs=1,
                             help="end date for the query period", required=True)
    args = args_parser.parse_args()
    file_path = args.file_path[0]

    begin = None
    if args.begin != None:
        begin = args.begin[0]
    end = None
    if args.end != None:
        end = args.end[0]

    if not is_datetime(begin) or not is_datetime(end):
        logger.error("The specified dates have an incorrect format")
        args_parser.print_help()
        sys.exit(-1)
    # end if
        
    generate_analysis(file_path, begin, end)

    logger.info("The analsys has been performed and saved into the file {}".format(file_path))
